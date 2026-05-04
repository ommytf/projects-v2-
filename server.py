#!/usr/bin/env python3
"""
MY ROUTINE • MY PURPOSE — Backend API Server
Flask + SQLite — Full database, exports, real-time support
"""
import sqlite3, json, os, io
from datetime import datetime, date, timedelta
from flask import Flask, request, jsonify, send_file, g
from flask_cors import CORS

app = Flask(__name__, static_folder='public', static_url_path='')
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'routine.db')

# ──────────────────────────────────────────────
# DATABASE SETUP
# ──────────────────────────────────────────────
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH, detect_types=sqlite3.PARSE_DECLTYPES)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
        g.db.execute("PRAGMA foreign_keys=ON")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db: db.close()

def init_db():
    with app.app_context():
        db = get_db()
        db.executescript("""
        CREATE TABLE IF NOT EXISTS settings (
            id      INTEGER PRIMARY KEY CHECK (id = 1),
            name    TEXT DEFAULT '',
            city    TEXT DEFAULT 'Kerege, Pemba',
            start_weight REAL DEFAULT 85,
            goal_weight  REAL DEFAULT 70,
            theme   TEXT DEFAULT 'dark',
            notifs  INTEGER DEFAULT 0,
            onboarded INTEGER DEFAULT 0,
            prayer_times TEXT DEFAULT '{"Fajr":"05:14","Dhuhr":"12:21","Asr":"15:43","Maghrib":"18:19","Isha":"19:25"}',
            habits_config TEXT DEFAULT '[]',
            schedule_config TEXT DEFAULT '[]',
            updated_at TEXT DEFAULT ''
        );
        INSERT OR IGNORE INTO settings (id) VALUES (1);

        CREATE TABLE IF NOT EXISTS daily_prayers (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            date     TEXT NOT NULL,
            prayer   TEXT NOT NULL,
            status   TEXT DEFAULT 'none',  -- none|prayed|missed|qada
            prayed_at TEXT,
            notes    TEXT DEFAULT '',
            UNIQUE(date, prayer)
        );

        CREATE TABLE IF NOT EXISTS daily_habits (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            habit_id TEXT NOT NULL,
            done    INTEGER DEFAULT 0,
            UNIQUE(date, habit_id)
        );

        CREATE TABLE IF NOT EXISTS weight_log (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            kg   REAL NOT NULL,
            belly_cm REAL,
            notes TEXT DEFAULT '',
            logged_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS daily_journal (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            date        TEXT NOT NULL UNIQUE,
            deen_rating TEXT,
            study_rate  INTEGER,
            energy_rate INTEGER,
            improve     TEXT DEFAULT '',
            grat1       TEXT DEFAULT '',
            grat2       TEXT DEFAULT '',
            grat3       TEXT DEFAULT '',
            journal     TEXT DEFAULT '',
            prayer_notes TEXT DEFAULT '',
            quran_juz   TEXT DEFAULT '',
            quran_pages INTEGER DEFAULT 0,
            quran_surah TEXT DEFAULT '',
            workout_type TEXT DEFAULT '',
            workout_notes TEXT DEFAULT '',
            soft_skill  TEXT DEFAULT '',
            day_score   INTEGER DEFAULT 0,
            updated_at  TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS study_log (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            subject TEXT NOT NULL,
            minutes INTEGER DEFAULT 0,
            logged_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS revision_log (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            content TEXT NOT NULL,
            logged_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS exercise_log (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            name    TEXT NOT NULL,
            sets    INTEGER DEFAULT 0,
            reps    INTEGER DEFAULT 0,
            notes   TEXT DEFAULT '',
            logged_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS meal_log (
            id      INTEGER PRIMARY KEY AUTOINCREMENT,
            date    TEXT NOT NULL,
            meal    TEXT NOT NULL,
            status  TEXT DEFAULT 'none',
            UNIQUE(date, meal)
        );

        CREATE TABLE IF NOT EXISTS pomodoro_log (
            id       INTEGER PRIMARY KEY AUTOINCREMENT,
            date     TEXT NOT NULL,
            duration INTEGER DEFAULT 25,
            completed INTEGER DEFAULT 1,
            logged_at TEXT DEFAULT ''
        );

        CREATE TABLE IF NOT EXISTS qada_log (
            id     INTEGER PRIMARY KEY AUTOINCREMENT,
            date   TEXT NOT NULL,
            prayer TEXT NOT NULL,
            count  INTEGER DEFAULT 0,
            UNIQUE(date, prayer)
        );

        CREATE INDEX IF NOT EXISTS idx_prayers_date ON daily_prayers(date);
        CREATE INDEX IF NOT EXISTS idx_habits_date  ON daily_habits(date);
        CREATE INDEX IF NOT EXISTS idx_journal_date ON daily_journal(date);
        CREATE INDEX IF NOT EXISTS idx_weight_date  ON weight_log(date);
        CREATE INDEX IF NOT EXISTS idx_study_date   ON study_log(date);
        """)
        db.commit()
        print("✓ Database initialised:", DB_PATH)

# ──────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────
def row_to_dict(row):
    return dict(row) if row else None

def rows_to_list(rows):
    return [dict(r) for r in rows]

def today_str():
    return date.today().isoformat()

def calc_day_score(date_str):
    db = get_db()
    prayers = db.execute("SELECT status FROM daily_prayers WHERE date=?", (date_str,)).fetchall()
    prayed = sum(1 for p in prayers if p['status']=='prayed')
    habits = db.execute("SELECT COUNT(*) as c FROM daily_habits WHERE date=? AND done=1", (date_str,)).fetchone()['c']
    journal = db.execute("SELECT journal, quran_pages, workout_type FROM daily_journal WHERE date=?", (date_str,)).fetchone()
    study = db.execute("SELECT COUNT(*) as c FROM study_log WHERE date=?", (date_str,)).fetchone()['c']
    score = 0
    score += (prayed / 5) * 40
    score += (habits / 14) * 30
    if journal:
        if journal['journal'] and len(journal['journal']) > 10: score += 10
        if journal['quran_pages'] and journal['quran_pages'] > 0: score += 10
        if journal['workout_type'] and journal['workout_type'] not in ('', 'Skip'): score += 5
    if study > 0: score += 5
    return round(score)

# ──────────────────────────────────────────────
# SETTINGS
# ──────────────────────────────────────────────
@app.route('/api/settings', methods=['GET'])
def get_settings():
    db = get_db()
    row = db.execute("SELECT * FROM settings WHERE id=1").fetchone()
    d = row_to_dict(row)
    if d:
        try: d['prayer_times'] = json.loads(d['prayer_times'])
        except: pass
    return jsonify(d)

@app.route('/api/settings', methods=['PUT'])
def update_settings():
    db = get_db()
    data = request.json or {}
    if 'prayer_times' in data and isinstance(data['prayer_times'], dict):
        data['prayer_times'] = json.dumps(data['prayer_times'])
    allowed = ['name','city','start_weight','goal_weight','theme','notifs','onboarded','prayer_times']
    updates = {k: v for k, v in data.items() if k in allowed}
    if updates:
        set_clause = ', '.join(f"{k}=?" for k in updates)
        db.execute(f"UPDATE settings SET {set_clause}, updated_at=datetime('now') WHERE id=1",
                   list(updates.values()))
        db.commit()
    row = db.execute("SELECT * FROM settings WHERE id=1").fetchone()
    d = row_to_dict(row)
    if d:
        try: d['prayer_times'] = json.loads(d['prayer_times'])
        except: pass
    return jsonify(d)

# ──────────────────────────────────────────────
# PRAYERS
# ──────────────────────────────────────────────
@app.route('/api/prayers/<date_str>', methods=['GET'])
def get_prayers(date_str):
    db = get_db()
    rows = db.execute("SELECT * FROM daily_prayers WHERE date=? ORDER BY id", (date_str,)).fetchall()
    prayers = rows_to_list(rows)
    qada = db.execute("SELECT * FROM qada_log WHERE date=?", (date_str,)).fetchall()
    return jsonify({'prayers': prayers, 'qada': rows_to_list(qada)})

@app.route('/api/prayers/<date_str>/<prayer>', methods=['PUT'])
def update_prayer(date_str, prayer):
    db = get_db()
    data = request.json or {}
    status = data.get('status', 'none')
    notes = data.get('notes', '')
    prayed_at = datetime.now().isoformat() if status == 'prayed' else None
    db.execute("""
        INSERT INTO daily_prayers (date, prayer, status, prayed_at, notes)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(date, prayer) DO UPDATE SET
            status=excluded.status, prayed_at=excluded.prayed_at, notes=excluded.notes
    """, (date_str, prayer, status, prayed_at, notes))
    db.commit()
    # Update day score
    score = calc_day_score(date_str)
    db.execute("INSERT INTO daily_journal (date, day_score) VALUES (?,?) ON CONFLICT(date) DO UPDATE SET day_score=?, updated_at=datetime('now')", (date_str, score, score))
    db.commit()
    return jsonify({'ok': True, 'score': score})

@app.route('/api/qada/<date_str>/<prayer>', methods=['PUT'])
def update_qada(date_str, prayer):
    db = get_db()
    count = request.json.get('count', 0)
    db.execute("INSERT INTO qada_log (date, prayer, count) VALUES (?,?,?) ON CONFLICT(date,prayer) DO UPDATE SET count=?",
               (date_str, prayer, count, count))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# HABITS
# ──────────────────────────────────────────────
@app.route('/api/habits/<date_str>', methods=['GET'])
def get_habits(date_str):
    db = get_db()
    rows = db.execute("SELECT * FROM daily_habits WHERE date=?", (date_str,)).fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/habits/<date_str>/<habit_id>', methods=['PUT'])
def update_habit(date_str, habit_id):
    db = get_db()
    done = 1 if request.json.get('done') else 0
    db.execute("INSERT INTO daily_habits (date, habit_id, done) VALUES (?,?,?) ON CONFLICT(date,habit_id) DO UPDATE SET done=?",
               (date_str, habit_id, done, done))
    db.commit()
    score = calc_day_score(date_str)
    db.execute("INSERT INTO daily_journal (date, day_score) VALUES (?,?) ON CONFLICT(date) DO UPDATE SET day_score=?, updated_at=datetime('now')", (date_str, score, score))
    db.commit()
    return jsonify({'ok': True, 'score': score})

# ──────────────────────────────────────────────
# JOURNAL / DAILY DATA
# ──────────────────────────────────────────────
@app.route('/api/journal/<date_str>', methods=['GET'])
def get_journal(date_str):
    db = get_db()
    row = db.execute("SELECT * FROM daily_journal WHERE date=?", (date_str,)).fetchone()
    study = db.execute("SELECT * FROM study_log WHERE date=? ORDER BY id", (date_str,)).fetchall()
    revision = db.execute("SELECT * FROM revision_log WHERE date=? ORDER BY id", (date_str,)).fetchall()
    exercises = db.execute("SELECT * FROM exercise_log WHERE date=? ORDER BY id", (date_str,)).fetchall()
    meals = db.execute("SELECT * FROM meal_log WHERE date=?", (date_str,)).fetchall()
    pomo = db.execute("SELECT COUNT(*) as c FROM pomodoro_log WHERE date=? AND completed=1", (date_str,)).fetchone()['c']
    return jsonify({
        'journal': row_to_dict(row),
        'study_log': rows_to_list(study),
        'revision_log': rows_to_list(revision),
        'exercises': rows_to_list(exercises),
        'meals': rows_to_list(meals),
        'pomo_today': pomo
    })

@app.route('/api/journal/<date_str>', methods=['PUT'])
def update_journal(date_str):
    db = get_db()
    data = request.json or {}
    allowed = ['deen_rating','study_rate','energy_rate','improve','grat1','grat2','grat3',
               'journal','prayer_notes','quran_juz','quran_pages','quran_surah',
               'workout_type','workout_notes','soft_skill']
    updates = {k: v for k, v in data.items() if k in allowed}
    # Recalculate score
    score = calc_day_score(date_str)
    updates['day_score'] = score
    updates['updated_at'] = datetime.now().isoformat()
    cols = list(updates.keys())
    vals = list(updates.values())
    placeholders = ','.join('?' for _ in cols)
    col_list = ','.join(cols)
    update_clause = ','.join(f"{c}=excluded.{c}" for c in cols)
    db.execute(f"INSERT INTO daily_journal (date,{col_list}) VALUES (?,{placeholders}) ON CONFLICT(date) DO UPDATE SET {update_clause}",
               [date_str] + vals)
    db.commit()
    return jsonify({'ok': True, 'score': score})

# ──────────────────────────────────────────────
# WEIGHT
# ──────────────────────────────────────────────
@app.route('/api/weight', methods=['GET'])
def get_weight():
    db = get_db()
    rows = db.execute("SELECT * FROM weight_log ORDER BY date DESC LIMIT 60").fetchall()
    return jsonify(rows_to_list(rows))

@app.route('/api/weight', methods=['POST'])
def log_weight():
    db = get_db()
    data = request.json or {}
    date_str = data.get('date', today_str())
    kg = float(data.get('kg', 0))
    belly = data.get('belly_cm')
    notes = data.get('notes', '')
    if kg < 30 or kg > 300:
        return jsonify({'error': 'Invalid weight'}), 400
    db.execute("INSERT INTO weight_log (date, kg, belly_cm, notes) VALUES (?,?,?,?) ON CONFLICT(date) DO UPDATE SET kg=?, belly_cm=?, notes=?, logged_at=datetime('now')",
               (date_str, kg, belly, notes, kg, belly, notes))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/weight/<date_str>', methods=['DELETE'])
def delete_weight(date_str):
    db = get_db()
    db.execute("DELETE FROM weight_log WHERE date=?", (date_str,))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# STUDY LOG
# ──────────────────────────────────────────────
@app.route('/api/study/<date_str>', methods=['POST'])
def add_study(date_str):
    db = get_db()
    data = request.json or {}
    db.execute("INSERT INTO study_log (date, subject, minutes) VALUES (?,?,?)",
               (date_str, data.get('subject',''), data.get('minutes', 25)))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/study/<int:entry_id>', methods=['DELETE'])
def delete_study(entry_id):
    db = get_db()
    db.execute("DELETE FROM study_log WHERE id=?", (entry_id,))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# EXERCISES
# ──────────────────────────────────────────────
@app.route('/api/exercises/<date_str>', methods=['POST'])
def add_exercise(date_str):
    db = get_db()
    data = request.json or {}
    db.execute("INSERT INTO exercise_log (date, name, sets, reps, notes) VALUES (?,?,?,?,?)",
               (date_str, data.get('name',''), data.get('sets',0), data.get('reps',0), data.get('notes','')))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/exercises/<int:entry_id>', methods=['DELETE'])
def delete_exercise(entry_id):
    db = get_db()
    db.execute("DELETE FROM exercise_log WHERE id=?", (entry_id,))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# REVISION LOG
# ──────────────────────────────────────────────
@app.route('/api/revision/<date_str>', methods=['POST'])
def add_revision(date_str):
    db = get_db()
    data = request.json or {}
    db.execute("INSERT INTO revision_log (date, content) VALUES (?,?)",
               (date_str, data.get('content','')))
    db.commit()
    return jsonify({'ok': True})

@app.route('/api/revision/<int:entry_id>', methods=['DELETE'])
def delete_revision(entry_id):
    db = get_db()
    db.execute("DELETE FROM revision_log WHERE id=?", (entry_id,))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# MEALS
# ──────────────────────────────────────────────
@app.route('/api/meals/<date_str>/<meal>', methods=['PUT'])
def update_meal(date_str, meal):
    db = get_db()
    status = request.json.get('status', 'none')
    db.execute("INSERT INTO meal_log (date, meal, status) VALUES (?,?,?) ON CONFLICT(date,meal) DO UPDATE SET status=?",
               (date_str, meal, status, status))
    db.commit()
    return jsonify({'ok': True})

# ──────────────────────────────────────────────
# POMODORO
# ──────────────────────────────────────────────
@app.route('/api/pomodoro', methods=['POST'])
def log_pomodoro():
    db = get_db()
    data = request.json or {}
    db.execute("INSERT INTO pomodoro_log (date, duration, completed) VALUES (?,?,?)",
               (today_str(), data.get('duration',25), 1))
    db.commit()
    total = db.execute("SELECT COUNT(*) as c FROM pomodoro_log WHERE completed=1").fetchone()['c']
    return jsonify({'ok': True, 'total': total})

# ──────────────────────────────────────────────
# STATS & PROGRESS
# ──────────────────────────────────────────────
@app.route('/api/stats', methods=['GET'])
def get_stats():
    db = get_db()
    today = today_str()

    # Salah streak
    streak = 0
    check = date.today()
    while True:
        ds = check.isoformat()
        rows = db.execute("SELECT COUNT(*) as c FROM daily_prayers WHERE date=? AND status='prayed'", (ds,)).fetchone()['c']
        if rows < 5: break
        streak += 1
        check -= timedelta(days=1)
        if streak > 365: break

    # Weight
    wl = db.execute("SELECT * FROM weight_log ORDER BY date DESC LIMIT 1").fetchone()
    settings = db.execute("SELECT start_weight, goal_weight FROM settings WHERE id=1").fetchone()
    sw = settings['start_weight'] if settings else 85
    gw = settings['goal_weight'] if settings else 70
    latest_kg = wl['kg'] if wl else sw
    lost = max(0, sw - latest_kg)
    pct_goal = round(lost / max(1, sw - gw) * 100)

    # Total sessions
    pomo_total = db.execute("SELECT COUNT(*) as c FROM pomodoro_log WHERE completed=1").fetchone()['c']

    # Weekly scores (last 5 weeks)
    weekly = []
    for w in range(4, -1, -1):
        week_start = date.today() - timedelta(days=date.today().weekday() + w*7)
        week_end = week_start + timedelta(days=6)
        rows = db.execute("SELECT day_score FROM daily_journal WHERE date >= ? AND date <= ?",
                           (week_start.isoformat(), week_end.isoformat())).fetchall()
        avg = round(sum(r['day_score'] for r in rows) / len(rows)) if rows else 0
        weekly.append({'week': str(week_start), 'score': avg, 'label': 'This week' if w==0 else f'{week_start.strftime("%b %d")}'})

    # 30-day heatmap
    heatmap = []
    for i in range(29, -1, -1):
        d = (date.today() - timedelta(days=i)).isoformat()
        p = db.execute("SELECT COUNT(*) as c FROM daily_prayers WHERE date=? AND status='prayed'", (d,)).fetchone()['c']
        heatmap.append({'date': d, 'prayed': p})

    # Total days tracked
    total_days = db.execute("SELECT COUNT(DISTINCT date) as c FROM daily_prayers WHERE status IN ('prayed','missed')").fetchone()['c']

    # Quran total pages
    quran_pages = db.execute("SELECT SUM(quran_pages) as s FROM daily_journal WHERE quran_pages > 0").fetchone()['s'] or 0
    quran_days = db.execute("SELECT COUNT(*) as c FROM daily_journal WHERE quran_pages > 0").fetchone()['c']

    # Journal days
    journal_days = db.execute("SELECT COUNT(*) as c FROM daily_journal WHERE length(journal) > 10").fetchone()['c']

    today_score = calc_day_score(today)

    return jsonify({
        'salah_streak': streak,
        'latest_kg': latest_kg,
        'lost_kg': round(lost, 1),
        'pct_goal': pct_goal,
        'pomo_total': pomo_total,
        'weekly_scores': weekly,
        'heatmap': heatmap,
        'total_days': total_days,
        'quran_pages': int(quran_pages),
        'quran_days': quran_days,
        'journal_days': journal_days,
        'today_score': today_score,
        'goal_weight': gw,
        'start_weight': sw,
    })

# ──────────────────────────────────────────────
# HISTORY / ALL DATA
# ──────────────────────────────────────────────
@app.route('/api/history', methods=['GET'])
def get_history():
    db = get_db()
    limit = int(request.args.get('limit', 90))
    rows = db.execute("""
        SELECT j.date, j.day_score, j.deen_rating, j.study_rate, j.energy_rate,
               j.quran_pages, j.workout_type, j.journal,
               (SELECT COUNT(*) FROM daily_prayers p WHERE p.date=j.date AND p.status='prayed') as prayed,
               (SELECT COUNT(*) FROM daily_habits h WHERE h.date=j.date AND h.done=1) as habits_done
        FROM daily_journal j ORDER BY j.date DESC LIMIT ?
    """, (limit,)).fetchall()
    return jsonify(rows_to_list(rows))

# ──────────────────────────────────────────────
# EXPORTS
# ──────────────────────────────────────────────
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    db = get_db()
    wb = Workbook()

    GOLD = "C9960C"
    NAVY = "050E1A"
    GREEN = "2A8C60"
    RED = "CC4444"
    LIGHT = "F5EFE0"

    def style_header(cell, bg=NAVY, fg=GOLD, bold=True):
        cell.font = Font(bold=bold, color=fg, name='Calibri')
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Daily Summary ──
    ws1 = wb.active
    ws1.title = "Daily Summary"
    ws1.sheet_view.showGridLines = False
    headers = ['Date','Day','Score','Prayers','Habits','Quran Pages','Workout','Deen Rating','Study Rate','Energy','Journal Preview']
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(1, c, h)
        style_header(cell)
        ws1.row_dimensions[1].height = 22
    rows = db.execute("""
        SELECT j.date, j.day_score, j.deen_rating, j.study_rate, j.energy_rate,
               j.quran_pages, j.workout_type, j.journal, j.updated_at,
               (SELECT COUNT(*) FROM daily_prayers p WHERE p.date=j.date AND p.status='prayed') as prayed,
               (SELECT COUNT(*) FROM daily_habits h WHERE h.date=j.date AND h.done=1) as habits_done
        FROM daily_journal j ORDER BY j.date DESC LIMIT 90
    """).fetchall()
    for r, row in enumerate(rows, 2):
        try: day_name = datetime.strptime(row['date'], '%Y-%m-%d').strftime('%A')
        except: day_name = ''
        vals = [row['date'], day_name, row['day_score'], f"{row['prayed']}/5",
                f"{row['habits_done']}/14", row['quran_pages'] or 0,
                row['workout_type'] or '', row['deen_rating'] or '',
                row['study_rate'] or '', row['energy_rate'] or '',
                (row['journal'] or '')[:60]]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(r, c, v)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(vertical='center')
            if c == 3:  # Score
                score = row['day_score'] or 0
                color = GREEN if score >= 80 else (GOLD if score >= 50 else RED)
                cell.fill = PatternFill("solid", fgColor=color)
                cell.font = Font(bold=True, color="FFFFFF", name='Calibri', size=10)
    auto_width(ws1)

    # ── Sheet 2: Prayer Log ──
    ws2 = wb.create_sheet("Prayer Log")
    ws2.sheet_view.showGridLines = False
    prayers = ['Fajr','Dhuhr','Asr','Maghrib','Isha']
    headers2 = ['Date'] + prayers + ['Total']
    for c, h in enumerate(headers2, 1):
        style_header(ws2.cell(1, c, h))
    prayer_rows = db.execute("SELECT date, prayer, status FROM daily_prayers ORDER BY date DESC, id").fetchall()
    date_map = {}
    for pr in prayer_rows:
        date_map.setdefault(pr['date'], {})[pr['prayer']] = pr['status']
    for r, (d, pm) in enumerate(sorted(date_map.items(), reverse=True)[:90], 2):
        ws2.cell(r, 1, d).font = Font(name='Calibri', size=10)
        total = 0
        for c, p in enumerate(prayers, 2):
            st = pm.get(p, 'none')
            cell = ws2.cell(r, c, '✔' if st=='prayed' else '✕' if st=='missed' else '—')
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Calibri', size=11, color=GREEN if st=='prayed' else (RED if st=='missed' else '888888'))
            if st == 'prayed': total += 1
        ws2.cell(r, 7, total).font = Font(bold=True, name='Calibri', size=10)
    auto_width(ws2)

    # ── Sheet 3: Habits Tracker ──
    ws3 = wb.create_sheet("Habits Tracker")
    ws3.sheet_view.showGridLines = False
    HABIT_LABELS = ['Fajr on Time','Quran 5+min','Qiyamu','3 Reality Checks','All 5 Salah','Dhikr 100+',
                    'Workout','No Rice/Junk','Water OK','No Soda','Ate Fruit','Study Recall','Wrapup Done','Sleep <23:00']
    HABIT_IDS = ['fajr','quran','qiyam','reality3','salah5','dhikr','workout','norice','water','nosoda','fruit','studyr','wrapup','sleep23']
    style_header(ws3.cell(1, 1, 'Habit'))
    dates_rows = db.execute("SELECT DISTINCT date FROM daily_habits ORDER BY date DESC LIMIT 30").fetchall()
    dates_list = [r['date'] for r in dates_rows]
    for c, d in enumerate(dates_list, 2):
        style_header(ws3.cell(1, c, d[-5:]))
    for r, (hid, hlbl) in enumerate(zip(HABIT_IDS, HABIT_LABELS), 2):
        ws3.cell(r, 1, hlbl).font = Font(name='Calibri', size=10)
        habit_data = db.execute("SELECT date, done FROM daily_habits WHERE habit_id=? AND date IN ({})".format(','.join('?'*len(dates_list))),
                                 [hid]+dates_list).fetchall()
        done_map = {row['date']: row['done'] for row in habit_data}
        for c, d in enumerate(dates_list, 2):
            done = done_map.get(d, 0)
            cell = ws3.cell(r, c, '✔' if done else '—')
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(name='Calibri', size=11, color=GREEN if done else '888888')
            if done: cell.fill = PatternFill("solid", fgColor="E8F8F0")
    auto_width(ws3)

    # ── Sheet 4: Weight Log ──
    ws4 = wb.create_sheet("Weight Log")
    ws4.sheet_view.showGridLines = False
    for c, h in enumerate(['Date','Weight (kg)','Belly (cm)','Change','Notes'], 1):
        style_header(ws4.cell(1, c, h))
    wrows = db.execute("SELECT * FROM weight_log ORDER BY date DESC").fetchall()
    prev_kg = None
    for r, wr in enumerate(wrows, 2):
        ws4.cell(r, 1, wr['date']).font = Font(name='Calibri', size=10)
        ws4.cell(r, 2, wr['kg']).font = Font(bold=True, name='Calibri', size=10)
        ws4.cell(r, 3, wr['belly_cm'] or '').font = Font(name='Calibri', size=10)
        if prev_kg:
            diff = round(wr['kg'] - prev_kg, 1)
            cell = ws4.cell(r, 4, f"{'+' if diff>0 else ''}{diff}")
            cell.font = Font(name='Calibri', size=10, color=GREEN if diff < 0 else RED)
        ws4.cell(r, 5, wr['notes'] or '').font = Font(name='Calibri', size=10)
        prev_kg = wr['kg']
    auto_width(ws4)

    # ── Sheet 5: Study Log ──
    ws5 = wb.create_sheet("Study Log")
    ws5.sheet_view.showGridLines = False
    for c, h in enumerate(['Date','Subject','Minutes','Logged At'], 1):
        style_header(ws5.cell(1, c, h))
    srows = db.execute("SELECT * FROM study_log ORDER BY date DESC, id DESC LIMIT 200").fetchall()
    for r, sr in enumerate(srows, 2):
        for c, v in enumerate([sr['date'], sr['subject'], sr['minutes'], sr['logged_at'][:16]], 1):
            ws5.cell(r, c, v).font = Font(name='Calibri', size=10)
    auto_width(ws5)

    # ── Sheet 6: Full Journal ──
    ws6 = wb.create_sheet("Journal")
    ws6.sheet_view.showGridLines = False
    for c, h in enumerate(['Date','Deen?','Study/10','Energy/10','Improve','Gratitude 1','Gratitude 2','Gratitude 3','Journal'], 1):
        style_header(ws6.cell(1, c, h))
    jrows = db.execute("SELECT * FROM daily_journal ORDER BY date DESC LIMIT 90").fetchall()
    for r, jr in enumerate(jrows, 2):
        for c, v in enumerate([jr['date'], jr['deen_rating'] or '', jr['study_rate'] or '', jr['energy_rate'] or '',
                                jr['improve'] or '', jr['grat1'] or '', jr['grat2'] or '', jr['grat3'] or '',
                                jr['journal'] or ''], 1):
            cell = ws6.cell(r, c, v)
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws6.column_dimensions['I'].width = 60
    auto_width(ws6)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"my-routine-{today_str()}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/api/export/json', methods=['GET'])
def export_json():
    db = get_db()
    data = {
        'exported_at': datetime.now().isoformat(),
        'settings': row_to_dict(db.execute("SELECT * FROM settings WHERE id=1").fetchone()),
        'prayers': rows_to_list(db.execute("SELECT * FROM daily_prayers ORDER BY date").fetchall()),
        'habits': rows_to_list(db.execute("SELECT * FROM daily_habits ORDER BY date").fetchall()),
        'weight_log': rows_to_list(db.execute("SELECT * FROM weight_log ORDER BY date").fetchall()),
        'journal': rows_to_list(db.execute("SELECT * FROM daily_journal ORDER BY date").fetchall()),
        'study_log': rows_to_list(db.execute("SELECT * FROM study_log ORDER BY date").fetchall()),
        'revision_log': rows_to_list(db.execute("SELECT * FROM revision_log ORDER BY date").fetchall()),
        'exercise_log': rows_to_list(db.execute("SELECT * FROM exercise_log ORDER BY date").fetchall()),
        'meal_log': rows_to_list(db.execute("SELECT * FROM meal_log ORDER BY date").fetchall()),
        'pomodoro_log': rows_to_list(db.execute("SELECT * FROM pomodoro_log ORDER BY date").fetchall()),
    }
    out = io.BytesIO(json.dumps(data, indent=2, default=str).encode())
    out.seek(0)
    return send_file(out, mimetype='application/json', as_attachment=True,
                     download_name=f"my-routine-backup-{today_str()}.json")

@app.route('/api/import/json', methods=['POST'])
def import_json():
    try:
        data = request.json
        db = get_db()
        if 'prayers' in data:
            for p in data['prayers']:
                db.execute("INSERT OR REPLACE INTO daily_prayers (date,prayer,status,prayed_at,notes) VALUES (?,?,?,?,?)",
                           (p.get('date'),p.get('prayer'),p.get('status','none'),p.get('prayed_at'),p.get('notes','')))
        if 'habits' in data:
            for h in data['habits']:
                db.execute("INSERT OR REPLACE INTO daily_habits (date,habit_id,done) VALUES (?,?,?)",
                           (h.get('date'),h.get('habit_id'),h.get('done',0)))
        if 'weight_log' in data:
            for w in data['weight_log']:
                db.execute("INSERT OR REPLACE INTO weight_log (date,kg,belly_cm,notes) VALUES (?,?,?,?)",
                           (w.get('date'),w.get('kg'),w.get('belly_cm'),w.get('notes','')))
        if 'journal' in data:
            for j in data['journal']:
                db.execute("INSERT OR REPLACE INTO daily_journal (date,journal,improve,deen_rating,study_rate,energy_rate,quran_pages,workout_type) VALUES (?,?,?,?,?,?,?,?)",
                           (j.get('date'),j.get('journal',''),j.get('improve',''),j.get('deen_rating'),j.get('study_rate'),j.get('energy_rate'),j.get('quran_pages',0),j.get('workout_type','')))
        db.commit()
        return jsonify({'ok': True, 'message': 'Data imported successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ──────────────────────────────────────────────
# SERVE FRONTEND
# ──────────────────────────────────────────────
@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/<path:path>')
def static_files(path):
    try:
        return app.send_static_file(path)
    except:
        return app.send_static_file('index.html')

# ──────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────
if __name__ == '__main__':
    os.makedirs('public', exist_ok=True)
    init_db()
    print("🕌 My Routine Server starting on http://localhost:5000")
    print("   Open this URL in your browser or phone on the same network")
    app.run(host='0.0.0.0', port=5000, debug=False)

@app.route('/api/reset', methods=['POST'])
def reset_data():
    db = get_db()
    tables = ['daily_prayers','daily_habits','weight_log','belly_log','daily_journal',
              'study_log','revision_log','exercise_log','meal_log','pomodoro_log','qada_log']
    for t in tables:
        try: db.execute(f"DELETE FROM {t}")
        except: pass
    db.execute("UPDATE settings SET onboarded=0 WHERE id=1")
    db.commit()
    return jsonify({'ok': True})
